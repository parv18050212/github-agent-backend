"""
Team Management Router - Phase 2
Handles all team CRUD operations and team-related endpoints
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, BackgroundTasks
from typing import List, Optional
from uuid import UUID, uuid4
from datetime import datetime
import csv
import io

from ..models import (
    Team, TeamCreate, TeamUpdate, TeamWithDetails,
    Student, StudentCreate,
    TeamList, PaginatedResponse
)
from ..schemas import (
    TeamResponse, TeamDetailResponse, TeamListResponse,
    TeamCreateRequest, TeamUpdateRequest,
    BulkUploadResponse, AnalysisJobResponse,
    MessageResponse, TeamAssignRequest, StudentGradeRequest
)
from ..middleware import get_current_user, RoleChecker, AuthUser
from ..database import get_supabase, get_supabase_admin_client

router = APIRouter(prefix="/api/teams", tags=["Teams"])


@router.get("", response_model=TeamListResponse)
async def list_teams(
    batch_id: Optional[UUID] = Query(None, description="Filter by batch ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
    mentor_id: Optional[UUID] = Query(None, description="Filter by assigned mentor"),
    search: Optional[str] = Query(None, description="Search by team name or repo URL"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    sort: str = Query("name", description="Sort field"),
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Get list of teams with filtering and pagination.
    
    - **Admin**: Can see all teams, requires batch_id parameter
    - **Mentor**: Only sees assigned teams
    """
    # Debug logging
    print(f"[Teams API] list_teams called")
    print(f"[Teams API] current_user.user_id: {current_user.user_id}")
    print(f"[Teams API] current_user.email: {current_user.email}")
    print(f"[Teams API] current_user.role: {current_user.role}")
    print(f"[Teams API] batch_id param: {batch_id}")
    
    supabase = get_supabase_admin_client()
    
    # Build query (base fields only; related data fetched separately for this page)
    query = supabase.table("teams").select(
        "id, team_name, batch_id, mentor_id, status, health_status, last_activity, "
        "created_at, updated_at, project_id, metadata",
        count="exact"
    )
    
    # Role-based filtering
    if current_user.role == "mentor":
        # Mentors only see their assigned teams
        # USE mentor_team_assignments table instead of teams.mentor_id
        print(f"[Teams API] Mentor detected: {current_user.user_id}")
        assignments = supabase.table("mentor_team_assignments")\
            .select("team_id")\
            .eq("mentor_id", str(current_user.user_id))\
            .execute()
        
        print(f"[Teams API] Assignments found: {len(assignments.data) if assignments.data else 0}")
        print(f"[Teams API] Assignment data: {assignments.data}")
        
        assigned_team_ids = [a["team_id"] for a in assignments.data] if assignments.data else []
        
        print(f"[Teams API] Assigned team IDs: {assigned_team_ids}")
        
        # If no assignments, return empty match (impossible ID) or handle gracefully
        if not assigned_team_ids:
             # Return empty list equivalent
             print("[Teams API] No assignments found, returning empty result")
             query = query.eq("id", "00000000-0000-0000-0000-000000000000") # Dummy UUID
        else:
             print(f"[Teams API] Filtering to {len(assigned_team_ids)} assigned teams")
             query = query.in_("id", assigned_team_ids)
            
    else:
        # Admins with is_mentor can see their own assigned teams when no filters provided
        if not batch_id and not mentor_id:
            try:
                mentor_flag = supabase.table("users").select("is_mentor").eq("id", str(current_user.user_id)).limit(1).execute()
                if mentor_flag.data and mentor_flag.data[0].get("is_mentor"):
                    assignments = supabase.table("mentor_team_assignments")\
                        .select("team_id")\
                        .eq("mentor_id", str(current_user.user_id))\
                        .execute()

                    assigned_team_ids = [a["team_id"] for a in assignments.data] if assignments.data else []
                    if not assigned_team_ids:
                        query = query.eq("id", "00000000-0000-0000-0000-000000000000")
                    else:
                        query = query.in_("id", assigned_team_ids)
                else:
                    raise HTTPException(
                        status_code=400,
                        detail="batch_id or mentor_id is required for admin users"
                    )
            except HTTPException:
                raise
            except Exception as e:
                print(f"[Teams API] Mentor flag lookup error: {e}")
                raise HTTPException(
                    status_code=400,
                    detail="batch_id or mentor_id is required for admin users"
                )
        else:
            # Admins: require either batch_id or mentor_id (for viewing mentor's teams)
            if mentor_id:
                # Admin viewing a specific mentor's teams - no batch_id required
                query = query.eq("mentor_id", str(mentor_id))
            elif batch_id:
                # Admin viewing teams in a batch
                query = query.eq("batch_id", str(batch_id))
            else:
                # If no filters provided, defaulting to empty or error?
                # Existing logic raised 400.
                raise HTTPException(
                    status_code=400,
                    detail="batch_id or mentor_id is required for admin users"
                )
    
    # Apply filters
    # Special handling for "unassigned" status - filter by mentor_id IS NULL
    if status == "unassigned":
        query = query.is_("mentor_id", "null")
    elif status:
        query = query.eq("status", status)
    
    # Apply mentor_id filter if batch_id was used (to filter within a batch)
    if batch_id and mentor_id:
        query = query.eq("mentor_id", str(mentor_id))
    
    if search:
        search_term = search.strip()
        matching_team_ids = set()

        team_search = supabase.table("teams").select("id").ilike("team_name", f"%{search_term}%").execute()
        for team in (team_search.data or []):
            if team.get("id"):
                matching_team_ids.add(team["id"])

        project_search = supabase.table("projects").select("team_id").ilike("repo_url", f"%{search_term}%").execute()
        for project in (project_search.data or []):
            if project.get("team_id"):
                matching_team_ids.add(project["team_id"])

        if not matching_team_ids:
            return TeamListResponse(
                teams=[],
                total=0,
                page=page,
                page_size=page_size,
                total_pages=0
            )

        query = query.in_("id", list(matching_team_ids))
    
    # Apply sorting (map frontend field names to database columns)
    sort_field = sort
    if sort == "name" or sort == "-name":
        # Frontend uses "name" but database has "team_name"
        sort_field = sort.replace("name", "team_name")
    
    if sort_field.startswith("-"):
        query = query.order(sort_field[1:], desc=True)
    else:
        query = query.order(sort_field)
    
    # Apply pagination
    offset = (page - 1) * page_size
    query = query.range(offset, offset + page_size - 1)
    
    # Execute query
    response = query.execute()

    teams = response.data or []
    total = response.count if hasattr(response, "count") else len(teams)

    project_ids = [team.get("project_id") for team in teams if team.get("project_id")]
    batch_ids = [team.get("batch_id") for team in teams if team.get("batch_id")]

    if batch_ids:
        batches_response = supabase.table("batches").select("id, name, semester, year").in_("id", batch_ids).execute()
        batch_map = {batch["id"]: batch for batch in (batches_response.data or [])}
    else:
        batch_map = {}

    if project_ids:
        projects_response = supabase.table("projects").select(
            "id, team_id, repo_url, total_score, status, last_analyzed_at"
        ).in_("id", project_ids).execute()
        project_map = {project["id"]: project for project in (projects_response.data or [])}
    else:
        project_map = {}

    for team in teams:
        batch = batch_map.get(team.get("batch_id"))
        if batch:
            team["batches"] = batch

        project = project_map.get(team.get("project_id"))
        if project:
            team["projects"] = [project]
            team["repo_url"] = project.get("repo_url")
        else:
            team["projects"] = []

    # Debug logging for query results
    print(f"[Teams API] Query returned {len(teams)} teams, total count: {total}")
    if current_user.role == "mentor":
        print(f"[Teams API] Mentor filter applied with mentor_id: {current_user.user_id}")

    # OPTIMIZED: Batch fetch mentor names (was: fetching after teams query)
    # Use cache to avoid repeated mentor queries
    from src.api.backend.utils.cache import cache, RedisCache
    
    mentor_ids = {str(team.get("mentor_id")) for team in teams if team.get("mentor_id")}
    
    if mentor_ids:
        # Try cache first (mentors change rarely)
        cache_key = "hackeval:mentors:lookup"
        mentor_lookup = cache.get(cache_key)
        
        if not mentor_lookup:
            # Cache miss - fetch all mentors and cache
            admin_supabase = get_supabase_admin_client()
            all_mentors = admin_supabase.table("users").select(
                "id, full_name, email, role, is_mentor"
            ).or_("role.eq.mentor,is_mentor.eq.true").execute()
            
            mentor_lookup = {
                str(m["id"]): m.get("full_name") or m.get("email") 
                for m in all_mentors.data or []
            }
            # Cache for 1 hour (mentors don't change often)
            cache.set(cache_key, mentor_lookup, RedisCache.TTL_LONG)
        
        # Assign mentor names to teams
        for team in teams:
            mentor_id = str(team.get("mentor_id")) if team.get("mentor_id") else None
            team["mentor_name"] = mentor_lookup.get(mentor_id) if mentor_id else None
    else:
        # No mentors to look up
        for team in teams:
            team["mentor_name"] = None
    
    return TeamListResponse(
        teams=teams,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size
    )


@router.post("", response_model=TeamResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def create_team(
    team_data: TeamCreateRequest,
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Create a new team (admin only).
    
    Creates team with students and optionally links to a project.
    """
    supabase = get_supabase_admin_client()
    
    # Verify batch exists
    batch_response = supabase.table("batches").select("id").eq("id", str(team_data.batch_id)).execute()
    if not batch_response.data:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    common_id = str(uuid4())
    
    # Create team
    team_insert = {
        "id": common_id,
        "batch_id": str(team_data.batch_id),
        "team_name": team_data.name,
        "health_status": "on_track",
        "student_count": len(team_data.students) if team_data.students else 0
    }
    
    try:
        team_response = supabase.table("teams").insert(team_insert).execute()
    except Exception as e:
        # Handle unique constraint for (batch_id, team_name)
        msg = str(e)
        if "unique_team_in_batch" in msg or "duplicate key" in msg:
            raise HTTPException(status_code=409, detail="A team with this name already exists in the selected batch")
        raise
    
    if not team_response.data:
        raise HTTPException(status_code=500, detail="Failed to create team")
    
    team = team_response.data[0]
    team_id = team["id"] # Should be common_id
    
    # Create project if repo URL provided
    if team_data.repo_url:
        project_insert = {
            "id": common_id,
            "team_id": team_id,
            "batch_id": str(team_data.batch_id),
            "team_name": team_data.name,
            "repo_url": team_data.repo_url,
            "status": "pending"
        }
        
        project_response = supabase.table("projects").insert(project_insert).execute()
        
        if project_response.data:
            # Update team with project_id
            supabase.table("teams").update({
                "project_id": common_id
            }).eq("id", team_id).execute()

            # Auto-queue analysis for newly created team
            try:
                from src.api.backend.crud import AnalysisJobCRUD, ProjectCRUD
                # Create analysis job
                job = AnalysisJobCRUD.create_job(UUID(common_id))
                job_id = job.get("id") if job else None

                # Update project status to queued
                ProjectCRUD.update_project_status(UUID(common_id), "queued")

                # Try to enqueue Celery task if available
                try:
                    from celery_worker import analyze_repository_task
                    task = analyze_repository_task.delay(
                        project_id=common_id,
                        job_id=str(job_id),
                        repo_url=team_data.repo_url,
                        team_name=team_data.name
                    )
                    supabase.table('analysis_jobs').update({
                        'metadata': {'celery_task_id': task.id}
                    }).eq('id', str(job_id)).execute()
                except Exception as celery_error:
                    print(f"⚠ Celery queueing failed for new team: {celery_error}")
            except Exception as analysis_error:
                print(f"⚠ Auto-analysis setup failed for new team: {analysis_error}")
    
    # Create students
    if team_data.students:
        students_insert = [
            {
                "team_id": team_id,
                "name": student.name,
                "email": student.email,
                "github_username": student.github_username
            }
            for student in team_data.students
        ]
        
        supabase.table("students").insert(students_insert).execute()
    
    # Fetch complete team data
    team_detail = supabase.table("teams").select(
        """
        *,
        batches(id, name, semester, year),
        students(*),
        projects!projects_teams_fk(*)
        """
    ).eq("id", team_id).execute()
    
    return TeamResponse(
        team=team_detail.data[0],
        message="Team created successfully"
    )


@router.delete("/clear-all", response_model=MessageResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def clear_all_teams(
    batch_id: UUID = Query(..., description="Batch ID to clear teams from"),
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Delete all teams and dependent data for a batch (admin only).
    """
    supabase = get_supabase_admin_client()

    teams_response = supabase.table("teams").select("id, project_id").eq("batch_id", str(batch_id)).execute()
    teams_data = teams_response.data or []

    if not teams_data:
        return MessageResponse(success=True, message="No teams found for this batch")

    team_ids = [team.get("id") for team in teams_data if team.get("id")]
    project_ids = {team.get("project_id") for team in teams_data if team.get("project_id")}

    projects_response = supabase.table("projects").select("id").eq("batch_id", str(batch_id)).execute()
    for project in projects_response.data or []:
        if project.get("id"):
            project_ids.add(project.get("id"))

    project_id_list = list(project_ids)

    if project_id_list:
        supabase.table("analysis_jobs").delete().in_("project_id", project_id_list).execute()
        supabase.table("analysis_snapshots").delete().in_("project_id", project_id_list).execute()
        supabase.table("issues").delete().in_("project_id", project_id_list).execute()
        supabase.table("project_comments").delete().in_("project_id", project_id_list).execute()
        supabase.table("tech_stack").delete().in_("project_id", project_id_list).execute()
        supabase.table("team_members").delete().in_("project_id", project_id_list).execute()

    if team_ids:
        supabase.table("mentor_team_assignments").delete().in_("team_id", team_ids).execute()
        supabase.table("students").delete().in_("team_id", team_ids).execute()

    supabase.table("teams").delete().eq("batch_id", str(batch_id)).execute()
    supabase.table("projects").delete().eq("batch_id", str(batch_id)).execute()

    return MessageResponse(success=True, message="Deleted all teams for this batch")


@router.post("/{team_id}/assign", response_model=MessageResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def assign_team_to_mentor(
    team_id: UUID,
    assignment: TeamAssignRequest,
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Assign a single team to a mentor (admin only).
    """
    supabase = get_supabase_admin_client()

    # Verify team exists
    team_response = supabase.table("teams").select("id, team_name, batch_id").eq("id", str(team_id)).execute()
    if not team_response.data:
        raise HTTPException(status_code=404, detail="Team not found")

    team = team_response.data[0]

    # Verify mentor exists
    mentor_response = supabase.table("users").select("id, email, full_name, role, is_mentor").eq(
        "id", str(assignment.mentor_id)
    ).execute()

    if not mentor_response.data:
        raise HTTPException(status_code=404, detail="Mentor not found")

    mentor = mentor_response.data[0]
    is_valid_mentor = mentor.get("role") == "mentor" or mentor.get("is_mentor") is True
    if not is_valid_mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")

    # Update team mentor_id
    supabase.table("teams").update({
        "mentor_id": str(assignment.mentor_id)
    }).eq("id", str(team_id)).execute()

    # Create assignment record if missing
    existing = supabase.table("mentor_team_assignments").select("id").eq(
        "mentor_id", str(assignment.mentor_id)
    ).eq("team_id", str(team_id)).execute()

    if not existing.data:
        supabase.table("mentor_team_assignments").insert({
            "mentor_id": str(assignment.mentor_id),
            "team_id": str(team_id),
            "batch_id": team.get("batch_id"),
            "assigned_by": str(current_user.user_id)
        }).execute()

    mentor_name = mentor.get("full_name") or mentor.get("email") or "mentor"

    return MessageResponse(
        success=True,
        message=f"Assigned {team.get('team_name', 'team')} to {mentor_name}"
    )


@router.post("/bulk-import", response_model=BulkUploadResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def bulk_import_teams_with_mentors(
    file: UploadFile = File(...),
    batch_id: UUID = Query(..., description="Batch ID for all teams"),
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Bulk import teams with mentor assignment via CSV or Excel (admin only).
    
    Supported formats: .csv, .xlsx
    
    CSV Format (Option A - with mentor):
    team_name,repo_url,mentor_email
    Team Alpha,https://github.com/org/alpha,drsmith@university.edu
    
    CSV Format (Option B - without mentor):
    team_name,repo_url
    Team Alpha,https://github.com/org/alpha
    
    Excel Format:
    Columns: Team Number, Team Name, Project Name, Github Link, etc.
    Maps: Team Name -> team_name, Github Link -> repo_url
    """
    supabase = get_supabase_admin_client()
    
    # Verify batch exists
    batch_response = supabase.table("batches").select("id, start_date").eq("id", str(batch_id)).execute()
    if not batch_response.data:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Determine file type and parse accordingly
    contents = await file.read()
    
    # We will parse into a structured list of teams to handle the "one row per student" format
    # structure: { "team_identifier": { "team_name": str, "repo_url": str, "mentor_email": str, "students": [ {name, email, ...} ] } }
    teams_map = {}
    
    if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
        # Parse Excel file
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        # Column Identification
        col_map = {}
        for i, header in enumerate(headers):
            h_lower = header.lower()
            if 'team no' in h_lower:
                col_map['team_id'] = i
            elif 'member name' in h_lower or 'name' in h_lower:
                col_map['student_name'] = i
            elif 'github' in h_lower or 'repo' in h_lower:
                col_map['repo_url'] = i
            elif 'mentor' in h_lower:
                col_map['mentor_email'] = i
            elif 'email' in h_lower:
                col_map['student_email'] = i
            elif 'roll' in h_lower:
                col_map['roll_no'] = i
            elif 'project' in h_lower and 'statement' in h_lower:
                col_map['project_desc'] = i
            elif 'section' in h_lower:
                col_map['section'] = i
            elif 'contact' in h_lower:
                col_map['contact'] = i
        
        # Check if we have the minimum required columns for the NEW format
        is_student_wise_format = 'team_id' in col_map and 'student_name' in col_map
        
        if is_student_wise_format:
            # === NEW FORMAT LOGIC (Student-wise rows with merged cells) ===
            current_team_no = None
            current_repo_url = None
            current_mentor = None
            current_project_desc = None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # helpers to safely get cell value
                def get_val(idx): return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else None
                
                # Check for Team No (Primary Key for grouping)
                raw_team_no = get_val(col_map.get('team_id'))
                
                # Forward Fill Logic
                if raw_team_no:
                    current_team_no = raw_team_no
                    # When specific team row starts, grab the other team-level attributes
                    current_repo_url = get_val(col_map.get('repo_url'))
                    current_mentor = get_val(col_map.get('mentor_email'))
                    current_project_desc = get_val(col_map.get('project_desc'))
                
                # If we don't have a team number yet (e.g. leading empty rows), skip
                if not current_team_no:
                    continue
                    
                # Normalize Team Name
                # If "Team No" is "1.0", make it "Team 1"
                try:
                    if current_team_no.endswith('.0'):
                        team_display_name = f"Team {int(float(current_team_no))}"
                    else:
                        team_display_name = f"Team {current_team_no}"
                except:
                    team_display_name = f"Team {current_team_no}"

                # Initialize team in map if needed
                if current_team_no not in teams_map:
                    teams_map[current_team_no] = {
                        "team_name": team_display_name,
                        "repo_url": current_repo_url or "",
                        "mentor_id": None, # Will lookup later
                        "mentor_email": current_mentor,
                        "project_description": current_project_desc or f"Project for {team_display_name}",
                        "team_no": current_team_no,
                        "project_statement": current_project_desc,
                        "mentor_raw": current_mentor,
                        "github_repository": current_repo_url,
                        "students": []
                    }
                
                # Add Student
                s_name = get_val(col_map.get('student_name'))
                if s_name:
                    teams_map[current_team_no]["students"].append({
                        "name": s_name,
                        "email": get_val(col_map.get('student_email')),
                        "roll_no": get_val(col_map.get('roll_no')),
                        "section": get_val(col_map.get('section')),
                        "contact": get_val(col_map.get('contact'))
                    })

        else:
            # === FALLBACK/OLD format logic (Team-wise rows) ===
            # Map column indices for old format
            team_name_col = None
            repo_url_col = None
            mentor_email_col = None
            student_emails_col = None
            
            for i, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    if 'team name' in header_lower or 'teamname' in header_lower:
                        team_name_col = i
                    elif 'github' in header_lower or 'repo' in header_lower or 'repository' in header_lower:
                        repo_url_col = i
                    elif 'mentor' in header_lower and 'email' in header_lower:
                        mentor_email_col = i
                    elif 'mail id' in header_lower or ('student' in header_lower and 'email' in header_lower):
                        student_emails_col = i

            if team_name_col is not None and repo_url_col is not None:
                 for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                     t_name = str(row[team_name_col]).strip() if row[team_name_col] else ""
                     if t_name:
                         t_id = f"row_{row_idx}" # temporary ID
                         teams_map[t_id] = {
                            "team_name": t_name,
                            "repo_url": str(row[repo_url_col]).strip() if row[repo_url_col] else "",
                            "mentor_email": str(row[mentor_email_col]).strip() if mentor_email_col is not None and row[mentor_email_col] else None,
                            "team_no": None,
                            "project_statement": None,
                            "mentor_raw": str(row[mentor_email_col]).strip() if mentor_email_col is not None and row[mentor_email_col] else None,
                            "github_repository": str(row[repo_url_col]).strip() if row[repo_url_col] else "",
                            "students": [] # Needs parsing from strings if present
                         }
                         # Old format parsing for students involved comma separated emails strings, 
                         # which we handle below in the loop mostly, but let's prep it here
                         stud_str = str(row[student_emails_col]).strip() if student_emails_col is not None and row[student_emails_col] else ""
                         if stud_str:
                             # Try to split by comma or newline
                             emails = [e.strip() for e in stud_str.replace('\n', ',').split(',') if e.strip()]
                             # We don't have names in this specific old format usually, or they were in team name
                             names = []
                             if '\n' in t_name:
                                 names = [n.strip() for n in t_name.split('\n') if n.strip()]
                             elif ',' in t_name:
                                 names = [n.strip() for n in t_name.split(',') if n.strip()]
                             else:
                                 names = [t_name] # One person team?
                             
                             # Zip them up best effort
                             for k in range(max(len(emails), len(names))):
                                 n = names[k] if k < len(names) else f"Student {k+1}"
                                 e = emails[k] if k < len(emails) else None
                                 teams_map[t_id]["students"].append({"name": n, "email": e})

    else:
        # Parse CSV file (Assuming old format for now as CSV doesn't support merge/grouping easily without strict schema)
        csv_file = io.StringIO(contents.decode("utf-8"))
        csv_reader = csv.DictReader(csv_file)
        for row_idx, row in enumerate(csv_reader):
              t_name = row.get("team_name", "").strip()
              if t_name:
                  t_id = f"csv_{row_idx}"
                  teams_map[t_id] = {
                      "team_name": t_name,
                      "repo_url": row.get("repo_url", "").strip(),
                      "mentor_email": row.get("mentor_email", "").strip(),
                      "team_no": None,
                      "project_statement": row.get("project_statement", "").strip() if row.get("project_statement") else None,
                      "mentor_raw": row.get("mentor_email", "").strip() if row.get("mentor_email") else None,
                      "github_repository": row.get("repo_url", "").strip(),
                      "students": []
                  }
                  # Parse students from CSV string if needed (similar to above)
                  # ... (Existing logic was simpler, just creating directly. We adopt structure now)
                  # For backward compat, we might just assume the user uses the Excel for the complex stuff.
                  # Simple recreation of students:
                  s_emails = row.get("student_emails", "")
                  if s_emails:
                       emails = [e.strip() for e in s_emails.replace('\n', ',').split(',') if e.strip()]
                       for e in emails:
                           teams_map[t_id]["students"].append({"name": "Student", "email": e})

    successful = 0
    failed = 0
    errors = []
    created_teams = []

    existing_team_rows = supabase.table("teams").select("id, team_name").eq(
        "batch_id", str(batch_id)
    ).execute()
    existing_team_map = {
        (row.get("team_name") or "").lower(): row
        for row in (existing_team_rows.data or [])
    }

    def _chunk(items, size=200):
        for i in range(0, len(items), size):
            yield items[i:i + size]

    teams_payload = []
    projects_payload = []
    students_payload = []
    
    # Get all mentors for email mapping
    mentors_response = supabase.table("users").select("id, email, full_name, role, is_mentor").or_("role.eq.mentor,is_mentor.eq.true").execute()
    # Create a map that tries to match by Email OR Name
    mentor_map = {}
    for m in mentors_response.data:
        if m.get("email"):
            mentor_map[m["email"].lower()] = m["id"]
        if m.get("full_name"):
            mentor_map[m["full_name"].lower()] = m["id"]

    # Fetch existing teams in this batch for idempotent imports
    existing_team_rows = supabase.table("teams").select("id, team_name, project_id").eq(
        "batch_id", str(batch_id)
    ).execute()
    existing_team_map = {
        (row.get("team_name") or "").lower(): row
        for row in (existing_team_rows.data or [])
    }

    def _chunk(items, size=200):
        for i in range(0, len(items), size):
            yield items[i:i + size]
    
    # Collect payloads for batch operations
    teams_payload = []
    projects_payload = []
    assignments_payload = []
    students_payload = []
    team_members_payload = []
    project_ids_for_members = []

    # Iterate over the Grouped Teams and Insert
    for team_key, team_data in teams_map.items():
        try:
            # Validate
            if not team_data["team_name"]:
                raise ValueError("Team name is missing")
            if not team_data["repo_url"]:
                # Some teams might not have a repo yet? 
                # Strict requirement: "Repository URL is required" in original code.
                # If we want to allow empty, change here. Assuming required for now.
                pass 

            # Resolve Mentor
            mentor_id = None
            m_input = team_data.get("mentor_email")
            if m_input:
                m_input = m_input.lower().strip()
                mentor_id = mentor_map.get(m_input)
                # Try partial match if not exact?
                if not mentor_id:
                     # fallback: try identifying by name from the map keys
                     for k, v in mentor_map.items():
                         if k in m_input or m_input in k:
                             mentor_id = v
                             break

            # Check if team already exists in batch
            existing_team = existing_team_map.get(team_data["team_name"].lower())
            is_existing = bool(existing_team)
            common_id = existing_team["id"] if is_existing else str(uuid4())
            
            # Prepare metadata with raw mentor info
            team_metadata = {}
            if team_data.get("mentor_email"):
                team_metadata["suggested_mentor"] = team_data.get("mentor_email")

            # Store raw ingestion columns for admin review and downstream use
            team_metadata["ingestion"] = {
                "team_no": team_data.get("team_no"),
                "project_statement": team_data.get("project_statement"),
                "mentor": team_data.get("mentor_raw") or team_data.get("mentor_email"),
                "github_repository": team_data.get("github_repository") or team_data.get("repo_url"),
                "students": [
                    {
                        "member_name": s.get("name"),
                        "roll_number": s.get("roll_no"),
                        "section": s.get("section"),
                        "email_id": s.get("email"),
                        "contact_number": s.get("contact")
                    }
                    for s in team_data.get("students", [])
                ]
            }

            team_id = common_id

            # Create or update Team
            team_payload = {
                "batch_id": str(batch_id),
                "team_name": team_data["team_name"],
                "repo_url": team_data.get("repo_url"),
                "mentor_id": mentor_id,
                "health_status": "on_track",
                "student_count": len(team_data["students"]),
                "metadata": team_metadata
            }
            if team_data.get("repo_url"):
                team_payload["project_id"] = team_id
            team_payload["id"] = team_id
            teams_payload.append(team_payload)
            
            # Create or update Project (project_id == team_id)
            if team_data.get("repo_url"):
                projects_payload.append({
                    "id": team_id,
                    "team_id": team_id,
                    "batch_id": str(batch_id),
                    "team_name": team_data["team_name"],
                    "repo_url": team_data["repo_url"],
                    "description": team_data.get("project_description", f"Project for {team_data['team_name']}"),
                    "status": "pending"
                })
            
            # Create Mentor Assignment
            if mentor_id:
                assignments_payload.append({
                    "mentor_id": mentor_id,
                    "team_id": team_id,
                    "batch_id": str(batch_id)
                })

            # Create Students
            students_to_insert = []
            team_members_to_insert = []  # For analytics
            
            for s in team_data["students"]:
                s_name = s["name"] or "Unknown Student"
                
                # Grading details / Metadata
                details = {}
                if s.get("roll_no"): details["roll_no"] = s.get("roll_no")
                if s.get("section"): details["section"] = s.get("section")
                if s.get("contact"): details["contact"] = s.get("contact")

                # Student Record
                students_to_insert.append({
                    "team_id": team_id,
                    "name": s_name,
                    "email": s.get("email"),
                    "grading_details": details
                })
                
                # Team Member Record (for analytics aggregation)
                if team_data.get("repo_url"): # Only if project exists
                    team_members_to_insert.append({
                        "project_id": common_id,
                        "name": s_name,
                        "commits": 0,
                        "contribution_pct": 0.0
                    })

            if students_to_insert:
                students_payload.extend(students_to_insert)

            if team_members_to_insert:
                team_members_payload.extend(team_members_to_insert)
                project_ids_for_members.append(team_id)

            created_teams.append({
                "id": team_id,
                "team_name": team_data.get("team_name"),
                "repo_url": team_data.get("repo_url"),
                "mentor_id": mentor_id,
                "batch_id": str(batch_id)
            })
            successful += 1

        except Exception as e:
            failed += 1
            errors.append({
                "row": team_key,
                "teamName": team_data.get("team_name", "Unknown"),
                "error": str(e)
            })

    # Batch write teams and projects
    try:
        if teams_payload:
            for chunk in _chunk(teams_payload, 200):
                supabase.table("teams").upsert(chunk, on_conflict="id").execute()

        if projects_payload:
            for chunk in _chunk(projects_payload, 200):
                supabase.table("projects").upsert(chunk, on_conflict="id").execute()

        if assignments_payload:
            # Filter out existing mentor assignments in batch
            team_ids = [a["team_id"] for a in assignments_payload]
            mentor_ids = [a["mentor_id"] for a in assignments_payload]
            existing_assignments = supabase.table("mentor_team_assignments").select("mentor_id, team_id").in_(
                "team_id", team_ids
            ).in_("mentor_id", mentor_ids).execute()
            existing_pairs = {
                (row.get("mentor_id"), row.get("team_id")) for row in (existing_assignments.data or [])
            }
            pending = [a for a in assignments_payload if (a["mentor_id"], a["team_id"]) not in existing_pairs]
            for chunk in _chunk(pending, 200):
                supabase.table("mentor_team_assignments").insert(chunk).execute()

        if students_payload:
            # Avoid merging distinct students that share an email by checking name mismatches
            emails = [s.get("email") for s in students_payload if s.get("email")]
            existing_email_map = {}
            if emails:
                existing = supabase.table("students").select("email, name").in_("email", emails).execute()
                existing_email_map = {
                    (row.get("email") or "").lower(): (row.get("name") or "").strip()
                    for row in (existing.data or [])
                }

            filtered_students = []
            for student in students_payload:
                email = (student.get("email") or "").strip()
                if not email:
                    filtered_students.append(student)
                    continue

                existing_name = existing_email_map.get(email.lower())
                new_name = (student.get("name") or "").strip()

                if existing_name and new_name and existing_name.lower() != new_name.lower():
                    errors.append({
                        "row": "bulk",
                        "teamName": student.get("team_id", "Unknown"),
                        "error": f"Student email already exists with different name: {email} ({existing_name} vs {new_name})"
                    })
                    continue

                filtered_students.append(student)

            for chunk in _chunk(filtered_students, 500):
                supabase.table("students").upsert(chunk, on_conflict="email").execute()

        if team_members_payload and project_ids_for_members:
            supabase.table("team_members").delete().in_("project_id", list(set(project_ids_for_members))).execute()
            for chunk in _chunk(team_members_payload, 500):
                supabase.table("team_members").insert(chunk).execute()
    except Exception as batch_error:
        failed += 1
        errors.append({
            "row": "bulk",
            "teamName": "batch",
            "error": f"Batch write failed: {str(batch_error)}"
        })

    return BulkUploadResponse(
        successful=successful,
        failed=failed,
        total=successful + failed,
        errors=errors,
        teams=created_teams, # Note: Returning partial objects here, might need schema adjustment if strict
        message=f"Bulk import completed: {successful} groups processed, {failed} failed"
    )


@router.post("/batch-upload", response_model=BulkUploadResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def bulk_upload_teams(
    file: UploadFile = File(...),
    batch_id: UUID = Query(..., description="Batch ID for all teams"),
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Bulk upload teams via CSV (admin only).
    
    CSV Format:
    teamName,repoUrl,description,student1Name,student1Email,student2Name,student2Email,...
    """
    supabase = get_supabase_admin_client()
    
    # Verify batch exists
    batch_response = supabase.table("batches").select("id").eq("id", str(batch_id)).execute()
    if not batch_response.data:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Read CSV
    contents = await file.read()
    csv_file = io.StringIO(contents.decode("utf-8"))
    csv_reader = csv.DictReader(csv_file)
    
    successful = 0
    failed = 0
    errors = []
    created_teams = []
    
    for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (header is row 1)
        try:
            team_name = row.get("teamName", "").strip()
            repo_url = row.get("repoUrl", "").strip()
            description = row.get("description", "").strip()
            
            if not team_name:
                raise ValueError("Team name is required")
            
            # Extract students (up to 6 students)
            students = []
            for i in range(1, 7):
                student_name = row.get(f"student{i}Name", "").strip()
                student_email = row.get(f"student{i}Email", "").strip()
                
                if student_name and student_email:
                    students.append({
                        "name": student_name,
                        "email": student_email
                    })
            
            existing_team = existing_team_map.get(team_name.lower())
            common_id = existing_team["id"] if existing_team else str(uuid4())

            team_payload = {
                "id": common_id,
                "batch_id": str(batch_id),
                "team_name": team_name,
                "health_status": "on_track",
                "student_count": len(students)
            }
            if repo_url:
                team_payload["project_id"] = common_id
            teams_payload.append(team_payload)
            team_id = common_id
            
            if repo_url:
                projects_payload.append({
                    "id": team_id,
                    "team_id": team_id,
                    "batch_id": str(batch_id),
                    "team_name": team_name,
                    "repo_url": repo_url,
                    "description": description,
                    "status": "pending"
                })
            
            if students:
                students_payload.extend([
                    {
                        "team_id": team_id,
                        "name": s["name"],
                        "email": s["email"]
                    }
                    for s in students
                ])
            
            created_teams.append({
                "id": team_id,
                "team_name": team_name,
                "repo_url": repo_url,
                "batch_id": str(batch_id)
            })
            successful += 1
            
        except Exception as e:
            failed += 1
            errors.append({
                "row": row_num,
                "teamName": row.get("teamName", ""),
                "error": str(e)
            })
    
    try:
        if teams_payload:
            for chunk in _chunk(teams_payload, 200):
                supabase.table("teams").upsert(chunk, on_conflict="id").execute()

        if projects_payload:
            for chunk in _chunk(projects_payload, 200):
                supabase.table("projects").upsert(chunk, on_conflict="id").execute()

        if students_payload:
            for chunk in _chunk(students_payload, 500):
                supabase.table("students").upsert(chunk, on_conflict="email").execute()
    except Exception as batch_error:
        failed += 1
        errors.append({
            "row": "bulk",
            "teamName": "batch",
            "error": f"Batch write failed: {str(batch_error)}"
        })

    return BulkUploadResponse(
        successful=successful,
        failed=failed,
        total=successful + failed,
        errors=errors,
        teams=created_teams,
        message=f"Batch upload completed: {successful} successful, {failed} failed"
    )


@router.get("/{team_id}")
async def get_team(
    team_id: UUID,
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Get detailed team information.
    
    Returns complete team data including students, project analysis, team members, and health status.
    Frontend expects: team_name, repo_url, batches, projects, team_members
    """
    supabase = get_supabase_admin_client()
    
    # Fetch team with all related data
    # Try by team ID first
    team_response = supabase.table("teams").select(
        """
        *,
        batches(id, name, semester, year),
        students(*),
        projects!projects_teams_fk(*)
        """
    ).eq("id", str(team_id)).execute()
    
    # If not found by ID, try by project_id (frontend may send project UUID)
    if not team_response.data:
        team_response = supabase.table("teams").select(
            """
            *,
            batches(id, name, semester, year),
            students(*),
            projects!projects_teams_fk(*)
            """
        ).eq("project_id", str(team_id)).execute()
    
    if not team_response.data:
        raise HTTPException(status_code=404, detail="Team not found")
    
    team = team_response.data[0]
    
    # DEBUG LOGGING
    print(f"[Teams API] get_team {team_id}")
    print(f"[Teams API] Team data: {team.keys()}")
    if "students" in team:
        print(f"[Teams API] Student count: {len(team['students'])}")
        print(f"[Teams API] Students: {team['students']}")
    else:
        print(f"[Teams API] 'students' key MISSING in response")
    
    # Check authorization
    if current_user.role == "mentor":
        mentor_id = str(current_user.user_id)
        if team.get("mentor_id") != mentor_id:
            assignment = supabase.table("mentor_team_assignments").select("id").eq(
                "mentor_id", mentor_id
            ).eq("team_id", str(team_id)).limit(1).execute()
            if not assignment.data:
                raise HTTPException(
                    status_code=403,
                    detail="Access denied: You are not assigned to this team"
                )
    
    # Fetch team members if project exists
    team_members = []
    project = team.get("projects")
    
    if project:
        project_id = project[0]["id"] if isinstance(project, list) else project["id"]
        members_response = supabase.table("team_members").select("*").eq("project_id", project_id).execute()
        team_members = members_response.data or []
    
    # Add team_members to response
    team["team_members"] = team_members
    
    # Ensure batches is always an object (not None)
    if not team.get("batches"):
        team["batches"] = {
            "id": team["batch_id"],
            "name": "Unknown Batch",
            "semester": "",
            "year": 0
        }
    
    return team


@router.put("/{team_id}", response_model=TeamResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def update_team(
    team_id: UUID,
    team_data: TeamUpdateRequest,
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Update team details (admin only).
    
    Allows updating team name, description, status, and health status.
    """
    supabase = get_supabase_admin_client()
    
    # Check if team exists
    existing_team = supabase.table("teams").select(
        "id, project_id, batch_id, team_name"
    ).eq("id", str(team_id)).execute()
    if not existing_team.data:
        raise HTTPException(status_code=404, detail="Team not found")
    team_row = existing_team.data[0]
    
    # Build update data
    update_data = {}
    if team_data.name is not None:
        update_data["team_name"] = team_data.name
    if team_data.status is not None:
        update_data["status"] = team_data.status
    if team_data.health_status is not None:
        update_data["health_status"] = team_data.health_status
    if team_data.risk_flags is not None:
        update_data["risk_flags"] = team_data.risk_flags
    if team_data.repo_url is not None:
        update_data["repo_url"] = team_data.repo_url
    
    # Update team
    team_response = supabase.table("teams").update(update_data).eq("id", str(team_id)).execute()

    # Sync or create project when repo_url changes
    if team_data.repo_url is not None:
        project_id = team_row.get("project_id")
        if project_id:
            supabase.table("projects").update({
                "repo_url": team_data.repo_url,
                "team_name": team_data.name or team_row.get("team_name")
            }).eq("id", str(project_id)).execute()
        elif team_data.repo_url:
            new_project_id = uuid4()
            project_insert = {
                "id": str(new_project_id),
                "team_id": str(team_id),
                "batch_id": str(team_row.get("batch_id")),
                "team_name": team_data.name or team_row.get("team_name"),
                "repo_url": team_data.repo_url,
                "status": "pending"
            }
            project_response = supabase.table("projects").insert(project_insert).execute()
            if project_response.data:
                supabase.table("teams").update({
                    "project_id": str(new_project_id)
                }).eq("id", str(team_id)).execute()
    
    # Update project if description provided
    if team_data.description is not None:
        supabase.table("projects").update({
            "description": team_data.description
        }).eq("team_id", str(team_id)).execute()
    
    # Fetch updated team
    updated_team = supabase.table("teams").select(
        """
        *,
        batches(id, name, semester, year),
        students(*),
        projects!projects_teams_fk(*)
        """
    ).eq("id", str(team_id)).execute()
    
    return TeamResponse(
        team=updated_team.data[0],
        message="Team updated successfully"
    )


@router.delete("/{team_id}", response_model=MessageResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def delete_team(
    team_id: UUID,
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Delete a team and its associated data (admin only).
    
    Cascades to delete students, project, and assignments.
    """
    supabase = get_supabase_admin_client()
    print(f"[Teams API] delete_team called for team_id={team_id}")
    
    # Check if team exists
    existing_team = supabase.table("teams").select("id, project_id").eq("id", str(team_id)).execute()
    if not existing_team.data:
        raise HTTPException(status_code=404, detail="Team not found")
    
    # Delete team (cascade will handle students and assignments)
    delete_response = supabase.table("teams").delete().eq("id", str(team_id)).execute()
    delete_error = getattr(delete_response, "error", None)
    if delete_error:
        raise HTTPException(status_code=500, detail=f"Failed to delete team: {delete_error}")
    
    return MessageResponse(
        success=True,
        message="Team deleted successfully"
    )


@router.get("/{team_id}/progress")
async def get_team_progress(
    team_id: UUID,
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Get weekly progress snapshots for a team.
    
    Returns historical analysis data showing improvement over time.
    """
    supabase = get_supabase()
    
    # Verify team exists and check authorization
    team_response = supabase.table("teams").select(
        "id, team_name, batch_id, mentor_id"
    ).eq("id", str(team_id)).execute()
    
    if not team_response.data:
        raise HTTPException(status_code=404, detail="Team not found")
    
    team = team_response.data[0]
    
    # Check authorization
    if current_user.role == "mentor":
        mentor_id = str(current_user.user_id)
        if team.get("mentor_id") != mentor_id:
            assignment = supabase.table("mentor_team_assignments").select("id").eq(
                "mentor_id", mentor_id
            ).eq("team_id", str(team_id)).limit(1).execute()
            if not assignment.data:
                raise HTTPException(
                    status_code=403,
                    detail="Access denied: You are not assigned to this team"
                )
    
    # Get all snapshots for this team
    snapshots_response = supabase.table("analysis_snapshots").select(
        """
        *,
        batch_analysis_runs(run_number, started_at, completed_at)
        """
    ).eq("team_id", str(team_id)).order("run_number", desc=False).execute()
    
    snapshots = snapshots_response.data or []
    
    # Calculate weekly trends
    weekly_data = []
    for snapshot in snapshots:
        run_info = snapshot.get("batch_analysis_runs", {})
        
        weekly_data.append({
            "week": snapshot["run_number"],
            "total_score": round(snapshot.get("total_score") or 0, 2),
            "originality_score": round(snapshot.get("originality_score") or 0, 2),
            "quality_score": round(snapshot.get("quality_score") or 0, 2),
            "security_score": round(snapshot.get("security_score") or 0, 2),
            "effort_score": round(snapshot.get("effort_score") or 0, 2),
            "commit_count": snapshot.get("commit_count", 0),
            "lines_of_code": snapshot.get("lines_of_code", 0),
            "analyzed_at": snapshot.get("analyzed_at"),
            "run_completed_at": run_info.get("completed_at") if isinstance(run_info, dict) else None
        })
    
    return weekly_data


@router.put("/{team_id}/grades", response_model=MessageResponse, dependencies=[Depends(RoleChecker(["admin", "mentor"]))])
async def update_student_grades(
    team_id: UUID,
    grades: List[StudentGradeRequest],
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Update grades for multiple students in a team.
    Allowed for Admins and the Assigned Mentor.
    """
    supabase = get_supabase_admin_client()
    
    # Verify team exists
    team_response = supabase.table("teams").select("id, mentor_id").eq("id", str(team_id)).execute()
    if not team_response.data:
        raise HTTPException(status_code=404, detail="Team not found")
    
    team = team_response.data[0]
    
    # Check permission
    if current_user.role == "mentor":
        mentor_id = str(current_user.user_id)
        if str(team.get("mentor_id")) != mentor_id:
            assignment = supabase.table("mentor_team_assignments").select("id").eq(
                "mentor_id", mentor_id
            ).eq("team_id", str(team_id)).limit(1).execute()
            if not assignment.data:
                raise HTTPException(status_code=403, detail="Not authorized to grade this team")
    
    # Verify all students belong to this team
    student_ids = [str(g.student_id) for g in grades]
    students_check = supabase.table("students").select("id").eq("team_id", str(team_id)).in_("id", student_ids).execute()
    
    found_ids = {s["id"] for s in students_check.data}
    if len(found_ids) != len(student_ids):
        raise HTTPException(status_code=400, detail="One or more students do not belong to this team")

    # Update grades (Supabase doesn't support bulk update easily with different values, so loop or upsert)
    # Upsert is better if we have all fields, but we only have grades. 
    # Looping is acceptable for small team sizes (~6 students).
    
    errors = []
    updated_count = 0
    
    for grade in grades:
        try:
            update_data = {}
            if current_user.role == "admin":
                # Admin grading
                if grade.admin_grade is not None:
                    update_data["admin_grade"] = grade.admin_grade
                if grade.admin_feedback is not None:
                    update_data["admin_feedback"] = grade.admin_feedback
            else:
                # Mentor grading
                if grade.grading_details:
                    current_student = supabase.table("students").select("grading_details").eq("id", str(grade.student_id)).execute()
                    if current_student.data:
                        current_details = current_student.data[0].get("grading_details") or {}
                        current_details.update(grade.grading_details)
                        update_data["grading_details"] = current_details
            
            
            if update_data:
                supabase.table("students").update(update_data).eq("id", str(grade.student_id)).execute()
                updated_count += 1
        except Exception as e:
            errors.append(f"Failed to update student {grade.student_id}: {str(e)}")
    
    if errors:
        return MessageResponse(success=False, message=f"Updated {updated_count} students with errors: {'; '.join(errors)}")
        
    return MessageResponse(success=True, message=f"Successfully graded {updated_count} students")

    # Calculate improvement from previous week
    improvement = 0
    if len(weekly_data) >= 2:
        latest = weekly_data[-1]
        previous = weekly_data[-2]
        improvement = latest["total_score"] - previous["total_score"]
    
    return {
        "team_id": str(team_id),
        "team_name": team["team_name"],
        "batch_id": team["batch_id"],
        "total_weeks": len(weekly_data),
        "current_score": weekly_data[-1]["total_score"] if weekly_data else 0,
        "improvement": round(improvement, 2),
        "weekly_data": weekly_data
    }


@router.post("/{team_id}/analyze", response_model=AnalysisJobResponse, dependencies=[Depends(RoleChecker(["admin"]))])
async def analyze_team(
    team_id: UUID,
    force: bool = Query(False, description="Force re-analysis (admin only)"),
    background_tasks: BackgroundTasks = None,
    current_user: AuthUser = Depends(get_current_user)
):
    """
    Trigger manual analysis for a team's repository (Admin only).
    
    Creates an analysis job and queues it for processing.
    
    Note: Only admins can manually trigger analysis. Mentors can view analysis
    results but cannot trigger new analyses. Automatic re-analysis is scheduled
    to run every 7 days via the batch analysis system.
    
    Use force=true to re-analyze immediately regardless of the last analysis time.
    """
    supabase = get_supabase()
    
    # Get team with project
    team_response = supabase.table("teams").select(
        "*, projects!projects_teams_fk(*)"
    ).eq("id", str(team_id)).execute()
    
    if not team_response.data:
        raise HTTPException(status_code=404, detail="Team not found")
    
    team = team_response.data[0]
    
    # Check if project exists
    project = team.get("projects")
    if not project:
        raise HTTPException(status_code=404, detail="Team has no associated project")
    
    # Handle both array and object formats
    project_data = project[0] if isinstance(project, list) else project
    project_id = project_data["id"]
    
    # For manual admin triggers, check if currently analyzing
    current_status = project_data.get("status")
    if current_status in ["analyzing", "queued"]:
        return AnalysisJobResponse(
            job_id=None,
            project_id=project_id,
            status="skipped",
            message=f"Analysis already in progress (status: {current_status})"
        )
    
    # Create analysis job
    job_insert = {
        "project_id": project_id,
        "status": "queued",
        "requested_by": str(current_user.user_id),
        "started_at": datetime.now().isoformat()
    }
    
    job_response = supabase.table("analysis_jobs").insert(job_insert).execute()
    
    if not job_response.data:
        raise HTTPException(status_code=500, detail="Failed to create analysis job")
    
    job = job_response.data[0]
    
    # Update project status
    supabase.table("projects").update({
        "status": "queued"
    }).eq("id", project_id).execute()
    
    # Queue task in the background to return fast
    def _queue_analysis_job():
        celery_queued = False
        try:
            from celery_worker import analyze_repository_task
            task = analyze_repository_task.delay(
                project_id=str(project_id),
                job_id=str(job["id"]),
                repo_url=project_data.get("repo_url"),
                team_name=team.get("team_name")
            )
            admin_supabase = get_supabase_admin_client()
            admin_supabase.table('analysis_jobs').update({
                'metadata': {'celery_task_id': task.id}
            }).eq('id', str(job["id"])).execute()
            celery_queued = True
        except Exception as celery_error:
            print(f"⚠ Celery queueing failed: {celery_error}")

        if not celery_queued:
            try:
                from src.api.backend.background import run_analysis_job
                run_analysis_job(
                    project_id=str(project_id),
                    job_id=str(job["id"]),
                    repo_url=project_data.get("repo_url"),
                    team_name=team.get("team_name")
                )
            except Exception as fallback_error:
                print(f"⚠ Fallback analysis failed: {fallback_error}")

    if background_tasks is not None:
        background_tasks.add_task(_queue_analysis_job)
    else:
        _queue_analysis_job()

    return AnalysisJobResponse(
        job_id=job["id"],
        project_id=project_id,
        status="queued",
        message="Analysis queued successfully"
    )
