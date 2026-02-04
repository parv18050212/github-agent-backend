"""
Test Excel import functionality
"""
import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('../top projects 3rd Year.xlsx')
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print('Headers found:', headers[:5])

# Map column indices
team_name_col = None
repo_url_col = None

for i, header in enumerate(headers):
    if header:
        header_lower = str(header).lower().strip()
        if 'team name' in header_lower:
            team_name_col = i
            print(f'✓ Team Name column: {i} ({header})')
        elif 'github' in header_lower:
            repo_url_col = i
            print(f'✓ GitHub Link column: {i} ({header})')

if team_name_col is None or repo_url_col is None:
    print('ERROR: Required columns not found!')
    exit(1)

print(f'\nParsing teams:')
teams = []
for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row[team_name_col] and row[repo_url_col]:
        team_name = str(row[team_name_col]).strip()
        repo_url = str(row[repo_url_col]).strip()
        
        # Clean up team name (remove newlines from multi-member names)
        team_name = team_name.replace('\n', ' ').replace('  ', ' ')
        
        teams.append({
            'team_name': team_name,
            'repo_url': repo_url
        })
        
        if len(teams) <= 3:
            print(f'  {len(teams)}. {team_name[:40]}...')
            print(f'     → {repo_url}')

print(f'\n✓ Total teams found: {len(teams)}')
print('✓ Excel parsing successful!')
